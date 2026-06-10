#!/usr/bin/env python3
"""Import Zernio analytics export zips into the avatar A/B test tracker.

Usage:
  python3 scripts/marketing/import_analytics.py                 # newest zip in the default exports folder
  python3 scripts/marketing/import_analytics.py export.zip      # one specific zip
  python3 scripts/marketing/import_analytics.py /dir/of/zips    # newest zip in that directory

Options:
  --tracker PATH   path to avatar_ab_test_tracker.xlsx (default: live copy in Google Drive)
  --map PATH       path to video_map.json (default: next to the tracker)
  --recalc         round-trip the tracker through LibreOffice to recalculate
                   formula columns (only if soffice is installed). Off by
                   default because the round-trip rewrites the whole file;
                   Excel / Numbers / Google Sheets recalc on open anyway.

What it writes, and nothing else:
  Sheet "Test Log" only. Columns C (Post Date), E (Views), H (Likes),
  I (Comments), J (Shares). Manual columns (D, F, G, K, L, M), formula
  columns (N-R), and all other sheets are never touched. Re-running with a
  newer export updates the metric cells (views grow over time, newer wins).

Matching (per the brief):
  1. video_map.json entry with a zernio_post_id -> match by Post ID.
  2. Fallback: same calendar day (America/New_York) between the export's
     Published Date and either the row's existing Post Date (C) or the
     video_map entry's posted_at. Only when exactly one TikTok post exists
     that day; ambiguous days are skipped and reported.

video_map.json is written at posting time by scripts/marketing/record_post.py.
NOTE for the future auto-posting step (Zernio TikTok publish, not built yet):
after a successful post it should write the same video_map.json entries that
record_post.py writes (video id -> zernio_post_id, tiktok_url, avatar_model,
posted_at), which makes matching here fully deterministic.
"""

import argparse
import csv
import json
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
import openpyxl

DATA_HOME = Path(
    "/Users/anderschan/Library/CloudStorage/GoogleDrive-dranders@drinzinna.com/"
    "My Drive/thepsychology.ai marketing/avatar-ab-test"
)
DEFAULT_TRACKER = DATA_HOME / "avatar_ab_test_tracker.xlsx"
DEFAULT_MAP = DATA_HOME / "video_map.json"
DEFAULT_EXPORTS_DIR = DATA_HOME / "exports"

LOCAL_TZ = ZoneInfo("America/New_York")
SHEET = "Test Log"
HEADER_ROW = 4
FIRST_DATA_ROW = 5
# The only cells this script is allowed to write. Everything else is manual
# (D, F, G, K, L, M), formulas (N-R), or other sheets.
AUTO_COLS = {"C": "Post Date", "E": "Views", "H": "Likes", "I": "Comments", "J": "Shares"}
METRIC_FIELDS = [("E", "Views"), ("H", "Likes"), ("I", "Comments"), ("J", "Shares")]
MANUAL_REMINDERS = [
    ("F", "Avg Watch Time (TikTok Studio)"),
    ("K", "Profile Visits (TikTok Studio)"),
    ("L", "Direct Traffic Spike (site analytics)"),
    ("M", "Signups (24h window)"),
]
REQUIRED_POST_COLS = ["Post ID", "Published Date", "Platform", "Post URL", "Likes", "Comments", "Shares", "Views"]


def fail(msg):
    print(f"ERROR: {msg}", file=sys.stderr)
    sys.exit(1)


def pick_zip(path):
    """Accept a zip path or a directory of zips; return the newest zip."""
    p = Path(path)
    if p.is_file():
        return p
    if not p.is_dir():
        fail(f"export path not found: {p}")
    zips = sorted(p.glob("*.zip"))
    if not zips:
        fail(f"no .zip files in {p}")

    def sort_key(z):
        m = re.search(r"(\d{4}-\d{2}-\d{2})", z.name)
        return (m.group(1) if m else "", z.stat().st_mtime)

    newest = max(zips, key=sort_key)
    if len(zips) > 1:
        print(f"{len(zips)} zips in {p}; using newest: {newest.name}")
    return newest


def find_csv(extract_dir, pattern):
    hits = sorted(Path(extract_dir).rglob(pattern))
    return hits[0] if hits else None


def local_day(iso_string):
    """ISO 8601 timestamp (with tz) -> calendar date in America/New_York."""
    dt = datetime.fromisoformat(str(iso_string).strip().replace("Z", "+00:00"))
    if dt.tzinfo is None:
        return dt.date()
    return dt.astimezone(LOCAL_TZ).date()


def cell_day(value):
    """Existing Post Date cell value -> date, or None."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.fromisoformat(str(value).strip()).date()
    except ValueError:
        print(f"WARNING: could not read existing Post Date value {value!r} as a date")
        return None


def as_int(value):
    """Export metric -> int, or None if missing/unreadable."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def load_video_map(path):
    p = Path(path)
    if not p.exists():
        print(f"WARNING: {p} not found. Post-ID matching disabled; date fallback only.")
        print("         Record posts with scripts/marketing/record_post.py to fix this.")
        return {}
    with open(p) as f:
        return json.load(f)


def load_posts(csv_path):
    df = pd.read_csv(csv_path, skipinitialspace=True)
    df.columns = [c.strip() for c in df.columns]
    missing = [c for c in REQUIRED_POST_COLS if c not in df.columns]
    if missing:
        fail(f"{csv_path.name} is missing expected columns {missing}; found {list(df.columns)}")
    tiktok = df[df["Platform"].astype(str).str.strip().str.lower() == "tiktok"].copy()
    return df, tiktok


def parse_followers(csv_path):
    """Parse the messy two-section followers CSV.

    Layout: a 'Current Follower Counts' section, then a 'Follower History
    (Daily)' section, each with its own header row (the second one sits
    mid-file). We detect header rows by their column names instead of
    trusting fixed line numbers.
    Returns (current_count, history) where history is [(date_str, count)],
    TikTok rows only. Either part may be None/empty if not found.
    """
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        rows = [[c.strip() for c in r] for r in csv.reader(f)]

    current = None
    history = []
    mode = None  # None | "current" | "history"
    cols = {}
    for row in rows:
        lowered = [c.lower() for c in row]
        if not any(lowered):
            continue
        if "platform" in lowered and "followers" in lowered:
            mode = "history" if "date" in lowered else "current"
            cols = {name: lowered.index(name) for name in ("platform", "followers", "date") if name in lowered}
            continue
        if mode is None or len(row) <= max(cols.values(), default=0):
            continue  # section title or stray line
        if row[cols["platform"]].lower() != "tiktok":
            continue
        try:
            count = int(float(row[cols["followers"]]))
        except ValueError:
            continue
        if mode == "current":
            current = count
        else:
            history.append((row[cols["date"]], count))
    history.sort()
    return current, history


def libreoffice_recalc(tracker):
    soffice = shutil.which("soffice")
    if not soffice:
        return "LibreOffice not found; Excel / Numbers will recalc formulas on open."
    with tempfile.TemporaryDirectory() as tmp:
        result = subprocess.run(
            [soffice, "--headless", "--convert-to", "xlsx", "--outdir", tmp, str(tracker)],
            capture_output=True, text=True,
        )
        out = Path(tmp) / tracker.name
        if result.returncode != 0 or not out.exists():
            return f"LibreOffice recalc failed ({result.stderr.strip() or result.returncode}); open in Excel to recalc."
        shutil.copy2(out, tracker)
    return "Recalculated formulas via LibreOffice."


def main():
    ap = argparse.ArgumentParser(description="Import a Zernio analytics export into the avatar A/B tracker.")
    ap.add_argument("export", nargs="?", default=str(DEFAULT_EXPORTS_DIR),
                    help=f"export zip, or a directory of them (default: {DEFAULT_EXPORTS_DIR})")
    ap.add_argument("--tracker", default=str(DEFAULT_TRACKER))
    ap.add_argument("--map", dest="map_path", default=str(DEFAULT_MAP))
    ap.add_argument("--recalc", action="store_true", help="recalculate formulas via LibreOffice if installed")
    args = ap.parse_args()

    tracker = Path(args.tracker)
    if not tracker.exists():
        fail(f"tracker not found: {tracker}")

    zip_path = pick_zip(args.export)
    print(f"== Avatar A/B import: {zip_path.name} ==")

    with tempfile.TemporaryDirectory() as tmp:
        try:
            with zipfile.ZipFile(zip_path) as z:
                z.extractall(tmp)
        except zipfile.BadZipFile:
            fail(f"{zip_path} is not a valid zip file")

        posts_csv = find_csv(tmp, "posts_analytics_*.csv")
        if not posts_csv:
            fail(f"no posts_analytics_*.csv inside {zip_path.name}")
        followers_csv = find_csv(tmp, "followers_*.csv")

        all_posts, tiktok = load_posts(posts_csv)
        print(f"Posts CSV: {posts_csv.name} ({len(all_posts)} rows, {len(tiktok)} tiktok)")

        followers_summary = None
        if followers_csv:
            try:
                current, history = parse_followers(followers_csv)
                if current is None and history:
                    current = history[-1][1]
                if current is not None:
                    if history:
                        first_date, first_count = history[0]
                        followers_summary = (f"TikTok followers: {current} now, "
                                             f"{current - first_count:+d} since {first_date} ({first_count})")
                    else:
                        followers_summary = f"TikTok followers: {current} now (no history rows found)"
                else:
                    followers_summary = f"WARNING: no TikTok rows found in {followers_csv.name}"
            except Exception as e:  # context only; never block the import
                followers_summary = f"WARNING: could not parse {followers_csv.name}: {e}"
        else:
            followers_summary = "WARNING: no followers_*.csv in the zip"

    # Index export rows for matching.
    by_post_id = {}
    by_day = defaultdict(list)
    for idx, row in tiktok.iterrows():
        by_post_id[str(row["Post ID"]).strip()] = idx
        by_day[local_day(row["Published Date"])].append(idx)

    video_map = load_video_map(args.map_path)

    wb = openpyxl.load_workbook(tracker)  # keeps formulas as formulas
    if SHEET not in wb.sheetnames:
        fail(f'sheet "{SHEET}" not found in {tracker} (has: {wb.sheetnames})')
    ws = wb[SHEET]

    updated, ambiguous, no_match_log, reminders = [], [], [], []
    consumed = set()

    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        vid = ws[f"A{r}"].value
        if vid is None or str(vid).strip() == "":
            continue
        vid = str(vid).strip()
        entry = video_map.get(vid, {})

        # 1) deterministic: Zernio post id from video_map.json
        match_idx, how = None, None
        post_id = (entry.get("zernio_post_id") or "").strip()
        if post_id and post_id in by_post_id:
            match_idx, how = by_post_id[post_id], f"post id {post_id[:8]}..."

        # 2) fallback: same calendar day, single TikTok post that day
        if match_idx is None:
            day = cell_day(ws[f"C{r}"].value)
            if day is None and entry.get("posted_at"):
                day = local_day(entry["posted_at"])
            if day is None:
                no_match_log.append(f"{vid}: no zernio_post_id match and no date to fall back on "
                                    "(fill Post Date in C, or record the post with record_post.py)")
            else:
                candidates = by_day.get(day, [])
                if len(candidates) == 1:
                    match_idx, how = candidates[0], f"date fallback {day}"
                elif len(candidates) == 0:
                    no_match_log.append(f"{vid}: no TikTok post in the export on {day}")
                else:
                    ambiguous.append(f"{vid}: {len(candidates)} TikTok posts on {day}, cannot pick one. "
                                     "Add its Zernio post id to video_map.json (record_post.py).")

        if match_idx is None:
            continue
        if match_idx in consumed:
            ambiguous.append(f"{vid}: matched the same export row as another video; skipped. Check video_map.json.")
            continue
        consumed.add(match_idx)
        export_row = tiktok.loc[match_idx]

        # Write auto cells only; record what changed.
        changes = []
        new_day = local_day(export_row["Published Date"])
        c_cell = ws[f"C{r}"]
        if cell_day(c_cell.value) != new_day:
            changes.append(("Post Date", c_cell.value, new_day.isoformat()))
            c_cell.value = new_day
            if c_cell.number_format == "General":
                c_cell.number_format = "yyyy-mm-dd"
        for col, field in METRIC_FIELDS:
            new = as_int(export_row[field])
            if new is None:
                print(f"WARNING: {vid}: export has no usable {field} value; left as is")
                continue
            cell = ws[f"{col}{r}"]
            if cell.value != new:
                changes.append((field, cell.value, new))
                cell.value = new

        label = f"{vid} ({how})"
        if changes:
            detail = ", ".join(f"{f} {old if old is not None else 'empty'} -> {new}" for f, old, new in changes)
            updated.append(f"{label}: {detail}")
        else:
            updated.append(f"{label}: already up to date")

        empty_manual = [name for col, name in MANUAL_REMINDERS
                        if ws[f"{col}{r}"].value in (None, "")]
        if empty_manual:
            reminders.append(f"{vid}: " + ", ".join(empty_manual))

    unmatched_export = []
    for idx, row in tiktok.iterrows():
        if idx in consumed:
            continue
        day = local_day(row["Published Date"])
        note = " (ambiguous day)" if len(by_day[day]) > 1 else ""
        unmatched_export.append(f'{day} {str(row["Post ID"]).strip()[:12]}... '
                                f'"{str(row["Content"])[:50]}"{note}')

    wrote = any(": " in u and "already up to date" not in u for u in updated)
    if wrote:
        backup = tracker.with_suffix(".backup.xlsx")
        shutil.copy2(tracker, backup)
        wb.save(tracker)
        print(f"Saved {tracker} (previous version kept at {backup.name})")
    else:
        print("No cell changes; tracker not rewritten.")

    print("\nMatched / updated:")
    for line in updated or ["  (none)"]:
        print(f"  {line}" if not line.startswith("  ") else line)
    if ambiguous:
        print("Skipped (ambiguous):")
        for line in ambiguous:
            print(f"  {line}")
    if no_match_log:
        print("Test Log rows with no export match:")
        for line in no_match_log:
            print(f"  {line}")
    if unmatched_export:
        print("Export rows with no Test Log match:")
        for line in unmatched_export:
            print(f"  {line}")
    if reminders:
        print("Manual fields still empty (grab from TikTok Studio / site analytics):")
        for line in reminders:
            print(f"  {line}")
    print(f"\n{followers_summary}")

    if wrote:
        if args.recalc:
            print(libreoffice_recalc(tracker))
        elif shutil.which("soffice"):
            print("Tip: pass --recalc to refresh formula columns via LibreOffice, or just open the file.")
        else:
            print("Formula columns N-R will recalculate when you open the file in Excel or Numbers.")


if __name__ == "__main__":
    main()
